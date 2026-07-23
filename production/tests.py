from django.test import TestCase

# Create your tests here.


import datetime as _dt
from django.contrib.auth.models import User as _User
from django.test import TestCase as _TestCase, Client as _Client
from django.urls import reverse as _reverse
from core.models import (
    WorkCenter as _WC, SubProcess as _SP, SubProcessType as _SPT, Shift as _Shift,
)
from users.models import UserProfile as _UP
from planning.models import (
    DailyPlan as _DP, HourlyPlan as _HP, Model as _PM,
)
from production.models import HourlyExecution as _HE


class ExecutionAjaxSaveTestCase(_TestCase):
    """The AJAX save path must UPDATE the single execution record (never
    duplicate), return the refreshed value, and report success — while still
    enforcing the comment/loss-reason rules on mismatches."""

    def setUp(self):
        self.user = _User.objects.create_user("ajaxleader", password="pw")
        _UP.objects.update_or_create(user=self.user, defaults={"role": "leader"})
        wc  = _WC.objects.create(name="WC-AJAX")
        spt = _SPT.objects.create(name="T-AJAX", applies_to="reactores", units_per_piece=1)
        sp  = _SP.objects.create(work_center=wc, name="SP-AJAX", subprocess_type=spt)
        shift = _Shift.objects.create(name="AJAX Shift", code="AJX",
                                      start_time=_dt.time(6), end_time=_dt.time(14),
                                      is_active=True)
        self.plan = _DP.objects.create(date=_dt.date(2026, 6, 20), work_center=wc,
                                       subprocess=sp, headcount=5, shift=shift)
        self.model = _PM.objects.create(name="AJAX Model")
        self.hp = _HP.objects.create(daily_plan=self.plan, hour=_dt.time(8),
                                     model=self.model, planned_quantity=10)
        self.client = _Client()
        self.client.login(username="ajaxleader", password="pw")
        self.url = _reverse("production:execution_enter", args=[self.plan.id])

    def _mgmt(self):
        pre = f"hp-{self.hp.id}-events"
        return {f"{pre}-TOTAL_FORMS": "0", f"{pre}-INITIAL_FORMS": "0",
                f"{pre}-MIN_NUM_FORMS": "0", f"{pre}-MAX_NUM_FORMS": "1000"}

    def _post(self, data):
        return self.client.post(self.url, {**data, **self._mgmt()},
                                HTTP_X_REQUESTED_WITH="XMLHttpRequest")

    def test_ajax_save_updates_not_duplicates(self):
        # Save on target (10 == 10): no comment needed → "Plan alcanzado".
        r1 = self._post({f"hp-{self.hp.id}-actual_quantity": "10",
                         f"hp-{self.hp.id}-scrap_quantity": "0"})
        d1 = r1.json()
        self.assertTrue(d1["ok"])
        self.assertEqual(_HE.objects.filter(hourly_plan=self.hp).count(), 1)

        # Save again over plan with the required comment → still ONE row.
        r2 = self._post({f"hp-{self.hp.id}-actual_quantity": "12",
                         f"hp-{self.hp.id}-scrap_quantity": "0",
                         f"hp-{self.hp.id}-over_comments": "Extra output"})
        d2 = r2.json()
        self.assertTrue(d2["ok"])
        self.assertEqual(_HE.objects.filter(hourly_plan=self.hp).count(), 1)
        self.assertEqual(_HE.objects.get(hourly_plan=self.hp).actual_quantity, 12)
        # Response echoes the refreshed value for the UI.
        saved = [row for row in d2["rows"] if row["saved"]][0]
        self.assertEqual(saved["actual_quantity"], 12)

    def test_ajax_missing_comment_returns_clear_error(self):
        # Below plan without loss reason / comment → must NOT save, and must
        # return a real error (not the misleading "enter a quantity" message).
        r = self._post({f"hp-{self.hp.id}-actual_quantity": "5",
                        f"hp-{self.hp.id}-scrap_quantity": "0"})
        d = r.json()
        self.assertFalse(d["ok"])
        self.assertEqual(_HE.objects.filter(hourly_plan=self.hp).count(), 0)
        row = [x for x in d["rows"] if not x["saved"]][0]
        self.assertTrue(row["errors"])

    def test_ajax_success_message_present(self):
        r = self._post({f"hp-{self.hp.id}-actual_quantity": "10",
                        f"hp-{self.hp.id}-scrap_quantity": "0"})
        d = r.json()
        self.assertEqual(d["level"], "success")
        self.assertIn("updated successfully", d["message"])


class ExecutionUpdateWorkflowTestCase(_TestCase):
    """Covers the update workflow: stale comments are replaced when the actual
    quantity changes, a fresh comment is mandatory on change (including the
    on-target case), and dependent calculations stay consistent."""

    def setUp(self):
        from production.models import LossReason
        self.user = _User.objects.create_user("upd", password="pw")
        _UP.objects.update_or_create(user=self.user, defaults={"role": "leader"})
        wc  = _WC.objects.create(name="WC-UPD")
        spt = _SPT.objects.create(name="T-UPD", applies_to="reactores", units_per_piece=1)
        sp  = _SP.objects.create(work_center=wc, name="SP-UPD", subprocess_type=spt)
        shift = _Shift.objects.create(name="UPD Shift", code="UPD",
                                      start_time=_dt.time(6), end_time=_dt.time(14),
                                      is_active=True)
        self.plan = _DP.objects.create(date=_dt.date(2026, 7, 5), work_center=wc,
                                       subprocess=sp, headcount=5, shift=shift)
        self.model = _PM.objects.create(name="UPD Model")
        self.hp = _HP.objects.create(daily_plan=self.plan, hour=_dt.time(8),
                                     model=self.model, planned_quantity=100)
        self.lr = LossReason.objects.create(name="Shortage")
        self.client = _Client()
        self.client.login(username="upd", password="pw")
        self.url = _reverse("production:execution_enter", args=[self.plan.id])

    def _mgmt(self):
        pre = f"hp-{self.hp.id}-events"
        return {f"{pre}-TOTAL_FORMS": "0", f"{pre}-INITIAL_FORMS": "0",
                f"{pre}-MIN_NUM_FORMS": "0", f"{pre}-MAX_NUM_FORMS": "1000"}

    def _post(self, data):
        return self.client.post(self.url, {**data, **self._mgmt()},
                                HTTP_X_REQUESTED_WITH="XMLHttpRequest")

    def test_decrease_replaces_stale_comment_and_recalculates(self):
        # Start on target (100 == 100).
        self._post({f"hp-{self.hp.id}-actual_quantity": "100",
                    f"hp-{self.hp.id}-scrap_quantity": "0"})
        e = _HE.objects.get(hourly_plan=self.hp)
        self.assertEqual(e.ok_comments, "Plan alcanzado")

        # Decrease to 80 with the required loss comment.
        self._post({f"hp-{self.hp.id}-actual_quantity": "80",
                    f"hp-{self.hp.id}-scrap_quantity": "0",
                    f"hp-{self.hp.id}-comments": "Material shortage",
                    f"hp-{self.hp.id}-loss_reasons": str(self.lr.id)})
        e.refresh_from_db()
        # New value saved, stale "Plan alcanzado" gone, loss comment stored.
        self.assertEqual(e.actual_quantity, 80)
        self.assertEqual(e.ok_comments, "")
        self.assertEqual(e.comments, "Material shortage")
        # Dependent calculations recalculated.
        self.assertEqual(e.diff_quantity, -20)
        self.assertEqual(e.efficiency_pct, 80.0)
        self.assertEqual(e.situation, _HE.SITUATION_BELOW)
        # No duplicate row.
        self.assertEqual(_HE.objects.filter(hourly_plan=self.hp).count(), 1)

    def test_change_without_comment_is_blocked(self):
        self._post({f"hp-{self.hp.id}-actual_quantity": "100",
                    f"hp-{self.hp.id}-scrap_quantity": "0"})
        # Try to change to 80 with no comment/loss reason → blocked, unchanged.
        r = self._post({f"hp-{self.hp.id}-actual_quantity": "80",
                        f"hp-{self.hp.id}-scrap_quantity": "0"})
        self.assertFalse(r.json()["ok"])
        e = _HE.objects.get(hourly_plan=self.hp)
        self.assertEqual(e.actual_quantity, 100)          # DB not modified
        self.assertEqual(e.ok_comments, "Plan alcanzado")  # no stale wipe either

    def test_on_target_change_autofills_plan_alcanzado(self):
        # First save below plan with a comment.
        self._post({f"hp-{self.hp.id}-actual_quantity": "80",
                    f"hp-{self.hp.id}-scrap_quantity": "0",
                    f"hp-{self.hp.id}-comments": "Late start",
                    f"hp-{self.hp.id}-loss_reasons": str(self.lr.id)})
        # Now change back to exactly 100 WITHOUT a comment → must SAVE and
        # auto-fill "Plan alcanzado", clearing the stale loss comment. No
        # comment is required on-target, even though the quantity changed.
        r = self._post({f"hp-{self.hp.id}-actual_quantity": "100",
                        f"hp-{self.hp.id}-scrap_quantity": "0"})
        self.assertTrue(r.json()["ok"])
        e = _HE.objects.get(hourly_plan=self.hp)
        self.assertEqual(e.actual_quantity, 100)
        self.assertEqual(e.ok_comments, "Plan alcanzado")
        self.assertEqual(e.comments, "")          # stale loss comment removed
        self.assertEqual(e.situation, _HE.SITUATION_OK)

        # If the user DOES type their own on-target note, it is respected.
        r = self._post({f"hp-{self.hp.id}-actual_quantity": "100",
                        f"hp-{self.hp.id}-scrap_quantity": "0",
                        f"hp-{self.hp.id}-ok_comments": "Recovered by end of hour"})
        self.assertTrue(r.json()["ok"])
        e.refresh_from_db()
        self.assertEqual(e.ok_comments, "Recovered by end of hour")

    def test_resave_same_value_not_forced_to_recomment(self):
        # On-target first capture, then re-save the SAME value with no comment.
        self._post({f"hp-{self.hp.id}-actual_quantity": "100",
                    f"hp-{self.hp.id}-scrap_quantity": "0"})
        r = self._post({f"hp-{self.hp.id}-actual_quantity": "100",
                        f"hp-{self.hp.id}-scrap_quantity": "0"})
        # Value did not change → no comment required, save succeeds.
        self.assertTrue(r.json()["ok"])


class ShiftOvertimeFilterTestCase(_TestCase):
    """Shift and Overtime filters on Execution and Hourly Plans, plus the
    Overtime column. Filters must compose without conflicts."""

    def setUp(self):
        import datetime as _d
        self.user = _User.objects.create_user("filteruser", password="pw")
        _UP.objects.update_or_create(user=self.user, defaults={"role": "leader"})
        self.wc = _WC.objects.create(name="WC-FILT")
        spt = _SPT.objects.create(name="T-FILT", applies_to="reactores", units_per_piece=1)
        self.sp = _SP.objects.create(work_center=self.wc, name="SP-FILT", subprocess_type=spt)
        self.shift_a = _Shift.objects.create(
            name="Shift A", code="FA",
            start_time=_d.time(6), end_time=_d.time(14), is_active=True)
        self.shift_b = _Shift.objects.create(
            name="Shift B", code="FB",
            start_time=_d.time(14), end_time=_d.time(22), is_active=True)
        self.model = _PM.objects.create(name="FILT Model")

        # Plan A: shift A, HAS overtime.
        self.plan_ot = _DP.objects.create(
            date=_d.date(2026, 8, 3), work_center=self.wc, subprocess=self.sp,
            headcount=5, shift=self.shift_a)
        _HP.objects.create(daily_plan=self.plan_ot, hour=_d.time(8),
                           model=self.model, planned_quantity=10, is_overtime=False)
        _HP.objects.create(daily_plan=self.plan_ot, hour=_d.time(15),
                           model=self.model, planned_quantity=5, is_overtime=True)

        # Plan B: shift B, NO overtime.
        self.plan_reg = _DP.objects.create(
            date=_d.date(2026, 8, 4), work_center=self.wc, subprocess=self.sp,
            headcount=5, shift=self.shift_b)
        _HP.objects.create(daily_plan=self.plan_reg, hour=_d.time(16),
                           model=self.model, planned_quantity=10, is_overtime=False)

        self.client = _Client()
        self.client.login(username="filteruser", password="pw")

    def _plan_ids(self, response):
        return {i["plan"].id for i in response.context["enriched"]}

    # ── Execution ────────────────────────────────────────────────────────
    def test_execution_overtime_only(self):
        r = self.client.get("/production/?overtime=1")
        ids = self._plan_ids(r)
        self.assertIn(self.plan_ot.id, ids)
        self.assertNotIn(self.plan_reg.id, ids)

    def test_execution_regular_only(self):
        r = self.client.get("/production/?overtime=0")
        ids = self._plan_ids(r)
        self.assertIn(self.plan_reg.id, ids)
        self.assertNotIn(self.plan_ot.id, ids)

    def test_execution_shift_filter(self):
        r = self.client.get(f"/production/?shift={self.shift_b.id}")
        ids = self._plan_ids(r)
        self.assertIn(self.plan_reg.id, ids)
        self.assertNotIn(self.plan_ot.id, ids)

    def test_execution_shift_and_overtime_compose(self):
        # Shift A + overtime only → just the OT plan.
        r = self.client.get(f"/production/?shift={self.shift_a.id}&overtime=1")
        self.assertEqual(self._plan_ids(r), {self.plan_ot.id})
        # Shift A + regular only → nothing (plan A has overtime).
        r2 = self.client.get(f"/production/?shift={self.shift_a.id}&overtime=0")
        self.assertNotIn(self.plan_ot.id, self._plan_ids(r2))

    def test_execution_has_overtime_flag(self):
        r = self.client.get("/production/")
        by_id = {i["plan"].id: i for i in r.context["enriched"]}
        self.assertTrue(by_id[self.plan_ot.id]["has_overtime"])
        self.assertFalse(by_id[self.plan_reg.id]["has_overtime"])
        self.assertIn("FILT Model", by_id[self.plan_ot.id]["ot_models"])

    # ── Hourly Plans ─────────────────────────────────────────────────────
    def test_hourly_board_overtime_filter(self):
        r = self.client.get("/hours/?overtime=1")
        ids = {c["plan"].id for c in r.context["board_cards"]}
        self.assertIn(self.plan_ot.id, ids)
        self.assertNotIn(self.plan_reg.id, ids)

    def test_hourly_board_shift_and_overtime_compose(self):
        r = self.client.get(f"/hours/?shift={self.shift_a.id}&overtime=1")
        ids = {c["plan"].id for c in r.context["board_cards"]}
        self.assertEqual(ids, {self.plan_ot.id})

    def test_no_filter_shows_all(self):
        r = self.client.get("/production/")
        ids = self._plan_ids(r)
        self.assertIn(self.plan_ot.id, ids)
        self.assertIn(self.plan_reg.id, ids)


class NonProductiveHourTestCase(_TestCase):
    """An hour planned at 0 is non-productive: the UI hides the input and the
    server clamps the actual to 0, so metrics can never be inflated."""

    def setUp(self):
        self.user = _User.objects.create_user("nonprod", password="pw")
        _UP.objects.filter(user=self.user).update(role="leader")
        self.user = _User.objects.get(pk=self.user.pk)
        wc = _WC.objects.create(name="WC-NP")
        spt = _SPT.objects.create(name="T-NP", applies_to="reactores",
                                  units_per_piece=1)
        sp = _SP.objects.create(work_center=wc, name="SP-NP",
                                subprocess_type=spt)
        shift = _Shift.objects.create(name="NP", code="NP",
                                      start_time=_dt.time(6),
                                      end_time=_dt.time(14), is_active=True)
        self.plan = _DP.objects.create(date=_dt.date(2026, 10, 25),
                                       work_center=wc, subprocess=sp,
                                       headcount=5, shift=shift)
        model = _PM.objects.create(name="NP Model")
        self.hp_zero = _HP.objects.create(daily_plan=self.plan,
                                          hour=_dt.time(8), model=model,
                                          planned_quantity=0)
        self.client = _Client()
        self.client.login(username="nonprod", password="pw")
        self.url = _reverse("production:execution_enter", args=[self.plan.id])

    def _mgmt(self, hp):
        pre = f"hp-{hp.id}-events"
        return {f"{pre}-TOTAL_FORMS": "0", f"{pre}-INITIAL_FORMS": "0",
                f"{pre}-MIN_NUM_FORMS": "0", f"{pre}-MAX_NUM_FORMS": "1000"}

    def test_actual_is_clamped_to_zero(self):
        """Even if a non-zero actual is submitted, it is stored as 0."""
        self.client.post(self.url, {
            f"hp-{self.hp_zero.id}-actual_quantity": "50",
            f"hp-{self.hp_zero.id}-scrap_quantity": "0",
            f"hp-{self.hp_zero.id}-zero_comment": "No production scheduled",
            **self._mgmt(self.hp_zero),
        }, HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        ex = _HE.objects.get(hourly_plan=self.hp_zero)
        self.assertEqual(ex.actual_quantity, 0)

    def test_no_editable_input_rendered(self):
        """The page must not offer an editable actual field for this hour."""
        html = self.client.get(self.url).content.decode()
        needle = f'name="hp-{self.hp_zero.id}-actual_quantity"'
        idx = html.find(needle)
        self.assertNotEqual(idx, -1)
        # Inspect the tag that carries this field name.
        start = html.rfind("<input", 0, idx)
        tag = html[start:idx + len(needle)]
        self.assertIn('type="hidden"', tag)


class SaveSummaryDetailTestCase(_TestCase):
    """Every row result — saved or failed — must carry the hour AND the model
    name, so the save summary can name exactly which rows need fixing."""

    def setUp(self):
        import random
        s = str(random.randint(10000, 99999))
        self.user = _User.objects.create_user("sumdet" + s, password="pw")
        _UP.objects.filter(user=self.user).update(role="leader")
        self.user = _User.objects.get(pk=self.user.pk)
        wc = _WC.objects.create(name="WC-SUM" + s)
        spt = _SPT.objects.create(name="T-SUM" + s, applies_to="reactores",
                                  units_per_piece=1)
        sp = _SP.objects.create(work_center=wc, name="SP-SUM" + s,
                                subprocess_type=spt)
        shift = _Shift.objects.create(name="SUM" + s, code="U" + s[:3],
                                      start_time=_dt.time(6),
                                      end_time=_dt.time(14), is_active=True)
        self.plan = _DP.objects.create(date=_dt.date(2026, 11, 15),
                                       work_center=wc, subprocess=sp,
                                       headcount=5, shift=shift)
        self.m_ok   = _PM.objects.create(name="OKMODEL" + s)
        self.m_bad  = _PM.objects.create(name="BADMODEL" + s)
        self.hp_ok  = _HP.objects.create(daily_plan=self.plan, hour=_dt.time(8),
                                         model=self.m_ok, planned_quantity=10)
        self.hp_bad = _HP.objects.create(daily_plan=self.plan, hour=_dt.time(9),
                                         model=self.m_bad, planned_quantity=10)
        self.client = _Client()
        self.client.login(username="sumdet" + s, password="pw")
        self.url = _reverse("production:execution_enter", args=[self.plan.id])

    def _mgmt(self, hp):
        pre = f"hp-{hp.id}-events"
        return {f"{pre}-TOTAL_FORMS": "0", f"{pre}-INITIAL_FORMS": "0",
                f"{pre}-MIN_NUM_FORMS": "0", f"{pre}-MAX_NUM_FORMS": "1000"}

    def test_failed_rows_include_hour_and_model(self):
        """A partial save must name the failing row, not just count it."""
        data = {
            **self._mgmt(self.hp_ok), **self._mgmt(self.hp_bad),
            f"hp-{self.hp_ok.id}-actual_quantity": "10",   # on target → saves
            f"hp-{self.hp_ok.id}-scrap_quantity": "0",
            f"hp-{self.hp_bad.id}-actual_quantity": "8",   # below plan, no reason
            f"hp-{self.hp_bad.id}-scrap_quantity": "0",
        }
        d = self.client.post(data=data, path=self.url,
                             HTTP_X_REQUESTED_WITH="XMLHttpRequest").json()
        saved  = [r for r in d["rows"] if r["saved"]]
        failed = [r for r in d["rows"] if not r["saved"]]

        self.assertEqual(len(saved), 1)
        self.assertEqual(len(failed), 1)
        # The failing row is fully identifiable.
        self.assertEqual(failed[0]["model"], self.m_bad.name)
        self.assertIn("09:00", failed[0]["hour"])
        self.assertTrue(failed[0]["errors"])
        # The saved row is identifiable too.
        self.assertEqual(saved[0]["model"], self.m_ok.name)

    def test_every_row_result_has_a_model(self):
        """No result may omit the model, whatever the outcome."""
        data = {
            **self._mgmt(self.hp_ok), **self._mgmt(self.hp_bad),
            f"hp-{self.hp_ok.id}-actual_quantity": "5",    # below plan, no reason
            f"hp-{self.hp_ok.id}-scrap_quantity": "0",
            f"hp-{self.hp_bad.id}-actual_quantity": "5",
            f"hp-{self.hp_bad.id}-scrap_quantity": "0",
        }
        d = self.client.post(data=data, path=self.url,
                             HTTP_X_REQUESTED_WITH="XMLHttpRequest").json()
        for row in d["rows"]:
            self.assertIn("model", row, f"row {row} is missing the model name")
            self.assertTrue(row["model"])


class HeadcountPlanVsActualTestCase(_TestCase):
    """Plan-level head count cascades to every hour, and Actuals captures how
    many people really worked, requiring an explanation when it differs."""

    def setUp(self):
        import random
        s = str(random.randint(10000, 99999))
        self.user = _User.objects.create_user("hcx" + s, password="pw")
        _UP.objects.filter(user=self.user).update(role="leader")
        self.user = _User.objects.get(pk=self.user.pk)
        wc = _WC.objects.create(name="WC-HCX" + s)
        spt = _SPT.objects.create(name="T-HCX" + s, applies_to="reactores",
                                  units_per_piece=1)
        sp = _SP.objects.create(work_center=wc, name="SP-HCX" + s,
                                subprocess_type=spt)
        shift = _Shift.objects.create(name="HCX" + s, code="K" + s[:3],
                                      start_time=_dt.time(6),
                                      end_time=_dt.time(14), is_active=True)
        self.plan = _DP.objects.create(date=_dt.date(2027, 7, 10),
                                       work_center=wc, subprocess=sp,
                                       headcount=5, shift=shift,
                                       created_by=self.user)
        m = _PM.objects.create(name="HCX Model" + s)
        self.hp_plain = _HP.objects.create(daily_plan=self.plan,
                                           hour=_dt.time(8), model=m,
                                           planned_quantity=10)
        self.hp_override = _HP.objects.create(daily_plan=self.plan,
                                              hour=_dt.time(9), model=m,
                                              planned_quantity=10, headcount=3)
        self.client = _Client()
        self.client.login(username="hcx" + s, password="pw")
        self.url = _reverse("production:execution_enter", args=[self.plan.id])

    def _mgmt(self, hp):
        pre = f"hp-{hp.id}-events"
        return {f"{pre}-TOTAL_FORMS": "0", f"{pre}-INITIAL_FORMS": "0",
                f"{pre}-MIN_NUM_FORMS": "0", f"{pre}-MAX_NUM_FORMS": "1000"}

    # ── Cascade from the plan ────────────────────────────────────────────
    def test_hours_without_override_follow_the_plan(self):
        from planning.services import update_headcount
        update_headcount(self.plan, 8, "more people", self.user)
        self.hp_plain.refresh_from_db()
        self.assertEqual(self.hp_plain.effective_headcount(), 8)

    def test_overrides_are_kept_by_default(self):
        from planning.services import update_headcount
        update_headcount(self.plan, 8, "more people", self.user)
        self.hp_override.refresh_from_db()
        self.assertEqual(self.hp_override.effective_headcount(), 3)

    def test_apply_to_all_clears_overrides(self):
        from planning.services import update_headcount
        ok, err, cleared = update_headcount(self.plan, 12, "everyone",
                                            self.user, apply_to_all=True)
        self.assertTrue(ok)
        self.assertEqual(cleared, 1)
        self.hp_override.refresh_from_db()
        self.assertEqual(self.hp_override.effective_headcount(), 12)

    # ── Capturing the real head count ────────────────────────────────────
    def test_difference_without_comment_is_blocked(self):
        r = self.client.post(self.url, {
            f"hp-{self.hp_plain.id}-actual_quantity": "10",
            f"hp-{self.hp_plain.id}-scrap_quantity": "0",
            f"hp-{self.hp_plain.id}-actual_headcount": "3",
            **self._mgmt(self.hp_plain),
        }, HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertFalse(r.json()["ok"])
        self.assertFalse(_HE.objects.filter(hourly_plan=self.hp_plain).exists())

    def test_difference_with_comment_saves(self):
        r = self.client.post(self.url, {
            f"hp-{self.hp_plain.id}-actual_quantity": "10",
            f"hp-{self.hp_plain.id}-scrap_quantity": "0",
            f"hp-{self.hp_plain.id}-actual_headcount": "3",
            f"hp-{self.hp_plain.id}-headcount_comment": "Two absences",
            **self._mgmt(self.hp_plain),
        }, HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertTrue(r.json()["ok"])
        ex = _HE.objects.get(hourly_plan=self.hp_plain)
        self.assertEqual(ex.actual_headcount, 3)
        self.assertEqual(ex.planned_headcount, 5)
        self.assertEqual(ex.headcount_diff, -2)

    def test_matching_headcount_needs_no_comment(self):
        r = self.client.post(self.url, {
            f"hp-{self.hp_plain.id}-actual_quantity": "10",
            f"hp-{self.hp_plain.id}-scrap_quantity": "0",
            f"hp-{self.hp_plain.id}-actual_headcount": "5",
            **self._mgmt(self.hp_plain),
        }, HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertTrue(r.json()["ok"])

    def test_blank_headcount_falls_back_to_plan(self):
        """Leaving it blank means 'same as planned' — never zero."""
        self.client.post(self.url, {
            f"hp-{self.hp_plain.id}-actual_quantity": "10",
            f"hp-{self.hp_plain.id}-scrap_quantity": "0",
            **self._mgmt(self.hp_plain),
        }, HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        ex = _HE.objects.get(hourly_plan=self.hp_plain)
        self.assertIsNone(ex.actual_headcount)
        self.assertEqual(ex.effective_actual_headcount, 5)
        self.assertEqual(ex.headcount_diff, 0)

    # ── Reports ──────────────────────────────────────────────────────────
    def test_exports_include_headcount_comparison(self):
        import io
        from openpyxl import load_workbook
        from analytics import day_service, export_service
        _HE.objects.create(hourly_plan=self.hp_plain, actual_quantity=10,
                           actual_headcount=3, headcount_comment="Absences")
        rep = day_service.build_day_report(self.plan)
        wb = load_workbook(io.BytesIO(export_service.day_report_to_excel(rep)))
        headers = [str(c.value) for row in wb.active.iter_rows()
                   for c in row if c.value]
        for col in ("HC Plan", "HC Actual", "HC Diff"):
            self.assertIn(col, headers)
        self.assertTrue(
            export_service.day_report_to_pdf(rep)[:5] == b"%PDF-")