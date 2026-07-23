import datetime as dt
from django.contrib.auth.models import User
from django.test import TestCase, Client

from core.models import WorkCenter, SubProcess, SubProcessType, Shift
from planning.models import DailyPlan, HourlyPlan, HourlyPlanBlock, Model as PlanningModel
from production.models import (
    HourlyExecution, ExecutionEvent, EventCategory, EventType,
)
from production.block_bridge import sync_event_for_block
from analytics import services


class MESArchitectureTestCase(TestCase):
    """End-to-end coverage of the MES refactor: automatic event generation
    from Time Blocks, duration auto-calc, category rollups, operator
    dimension, and the KPI service layer."""

    def setUp(self):
        self.op = User.objects.create_user("operator1", password="pw")
        self.wc = WorkCenter.objects.create(name="WC-MES")
        spt = SubProcessType.objects.create(
            name="T-MES", applies_to="reactores", units_per_piece=1)
        self.sp = SubProcess.objects.create(
            work_center=self.wc, name="SP-MES", subprocess_type=spt)
        self.shift = Shift.objects.create(
            name="MES Shift", code="MES",
            start_time=dt.time(6, 0), end_time=dt.time(14, 0), is_active=True)
        self.plan = DailyPlan.objects.create(
            date=dt.date(2026, 6, 10), work_center=self.wc, subprocess=self.sp,
            headcount=5, shift=self.shift, operator=self.op)
        self.model = PlanningModel.objects.create(name="MES Model")
        self.hp = HourlyPlan.objects.create(
            daily_plan=self.plan, hour=dt.time(8, 0),
            model=self.model, planned_quantity=100)

    # ── Categories seeded ────────────────────────────────────────────────
    def test_categories_are_seeded(self):
        codes = set(EventCategory.objects.values_list("code", flat=True))
        self.assertTrue({"lunch", "preop", "workfin", "chair", "extra"} <= codes)

    def test_event_types_link_to_categories(self):
        lunch = EventType.objects.get(name="Lunch")
        self.assertIsNotNone(lunch.category)
        self.assertEqual(lunch.category.code, "lunch")

    # ── Auto event generation from a Time Block ──────────────────────────
    def test_saving_block_auto_creates_event(self):
        block = HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(8, 0),
            block_type="lunch", minutes=30, created_by=self.op)
        ev = ExecutionEvent.objects.get(source_block=block)
        self.assertEqual(ev.duration_minutes, 30)
        self.assertEqual(ev.source, ExecutionEvent.SOURCE_BLOCK)
        self.assertEqual(ev.event_type.category.code, "lunch")

    def test_block_event_generation_is_idempotent(self):
        block = HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(8, 0),
            block_type="preop", minutes=15)
        block.minutes = 20
        block.save()
        block.save()
        events = ExecutionEvent.objects.filter(source_block=block)
        self.assertEqual(events.count(), 1)
        self.assertEqual(events.first().duration_minutes, 20)

    def test_deleting_block_removes_its_event(self):
        block = HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(8, 0),
            block_type="chair", minutes=5)
        self.assertEqual(ExecutionEvent.objects.filter(source_block=block).count(), 1)
        block.delete()
        self.assertEqual(ExecutionEvent.objects.filter(source="block").count(), 0)

    # ── Duration auto-calculation ─────────────────────────────────────────
    def test_duration_derived_from_start_end(self):
        execution, _ = HourlyExecution.objects.get_or_create(hourly_plan=self.hp)
        other = EventType.objects.get(name="Other")
        ev = ExecutionEvent.objects.create(
            execution=execution, event_type=other, duration_minutes=0,
            start_time=dt.time(8, 0), end_time=dt.time(8, 25))
        self.assertEqual(ev.duration_minutes, 25)

    # ── Plan achieved ─────────────────────────────────────────────────────
    def test_plan_achieved_comment_when_equal(self):
        execution = HourlyExecution(hourly_plan=self.hp, actual_quantity=100)
        self.assertTrue(execution.is_plan_achieved)
        self.assertEqual(execution.resolve_ok_comment(""), "Plan alcanzado")

    def test_user_comment_not_overwritten(self):
        execution = HourlyExecution(hourly_plan=self.hp, actual_quantity=100)
        self.assertEqual(execution.resolve_ok_comment("Great hour"), "Great hour")

    # ── KPI service ───────────────────────────────────────────────────────
    def test_kpi_summary_reflects_events(self):
        HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(8, 0),
            block_type="lunch", minutes=30)
        HourlyExecution.objects.filter(hourly_plan=self.hp).update(actual_quantity=100)
        k = services.kpi_summary(work_center_id=self.wc.id)
        self.assertEqual(k["downtime_minutes"], 30)
        self.assertEqual(k["planned_downtime_minutes"], 30)   # lunch is planned
        self.assertEqual(k["number_of_events"], 1)

    def test_by_operator_dimension(self):
        HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(8, 0),
            block_type="lunch", minutes=30)
        rows = services.by_operator(work_center_id=self.wc.id)
        names = [r["execution__hourly_plan__daily_plan__operator__username"] for r in rows]
        self.assertIn("operator1", names)

    def test_planned_vs_actual_and_achievement(self):
        HourlyExecution.objects.filter(hourly_plan=self.hp).delete()
        HourlyExecution.objects.create(hourly_plan=self.hp, actual_quantity=90)
        pva = services.planned_vs_actual(work_center_id=self.wc.id)
        self.assertEqual(pva["planned"], 100)
        self.assertEqual(pva["actual"], 90)
        self.assertEqual(services.plan_achievement_pct(work_center_id=self.wc.id), 90.0)


class BlockEventSyncProtectionTestCase(TestCase):
    """Time Block → Event sync must seed the event once with the same title
    and time, keep it editable, and NEVER overwrite a user's edits when the
    block is re-saved."""

    def setUp(self):
        self.u = User.objects.create_user("syncuser", password="pw")
        self.wc = WorkCenter.objects.create(name="WC-SYNC2")
        spt = SubProcessType.objects.create(
            name="T-SYNC2", applies_to="reactores", units_per_piece=1)
        self.sp = SubProcess.objects.create(
            work_center=self.wc, name="SP-SYNC2", subprocess_type=spt)
        self.shift = Shift.objects.create(
            name="SYNC2", code="SN2",
            start_time=dt.time(6), end_time=dt.time(14), is_active=True)
        self.plan = DailyPlan.objects.create(
            date=dt.date(2026, 7, 20), work_center=self.wc, subprocess=self.sp,
            headcount=5, shift=self.shift, operator=self.u)
        self.model = PlanningModel.objects.create(name="SYNC2 Model")
        self.hp = HourlyPlan.objects.create(
            daily_plan=self.plan, hour=dt.time(12),
            model=self.model, planned_quantity=100)

    def test_block_seeds_event_with_same_title_and_time(self):
        block = HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(12),
            block_type="lunch", minutes=30, created_by=self.u)
        ev = ExecutionEvent.objects.get(source_block=block)
        self.assertEqual(ev.event_type.name, "Lunch")   # same title
        self.assertEqual(ev.duration_minutes, 30)        # same time
        self.assertEqual(ev.source, ExecutionEvent.SOURCE_BLOCK)

    def test_user_edit_survives_block_resave(self):
        block = HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(12),
            block_type="lunch", minutes=30, created_by=self.u)
        ev = ExecutionEvent.objects.get(source_block=block)

        # User edits the event and it is flagged as a human edit.
        ev.duration_minutes = 25
        ev.comment = "Only took 25 min"
        ev.user_modified = True
        ev.save()

        # Re-saving the block must NOT overwrite the user's edit.
        block.save()
        ev.refresh_from_db()
        self.assertEqual(ev.duration_minutes, 25)
        self.assertEqual(ev.comment, "Only took 25 min")

    def test_unedited_event_still_syncs_with_block(self):
        # If the user never edited it, the block remains the source and can
        # still update the event (e.g. plan corrected from 30 to 20 min).
        block = HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(12),
            block_type="lunch", minutes=30, created_by=self.u)
        block.minutes = 20
        block.save()
        ev = ExecutionEvent.objects.get(source_block=block)
        self.assertEqual(ev.duration_minutes, 20)   # synced, not user-locked


class DayDashboardTestCase(TestCase):
    """Phase 1 DAY dashboard: unit→piece conversion, classification, totals,
    completion %, and read-only rendering."""

    def setUp(self):
        self.u = User.objects.create_user("dayuser", password="pw")
        self.wc = WorkCenter.objects.create(name="WC-DAY")
        # 3 units = 1 piece — the critical conversion case.
        spt = SubProcessType.objects.create(
            name="T-DAY", applies_to="reactores", units_per_piece=3)
        self.sp = SubProcess.objects.create(
            work_center=self.wc, name="SP-DAY", subprocess_type=spt)
        self.shift = Shift.objects.create(
            name="DAY", code="DAY",
            start_time=dt.time(6), end_time=dt.time(14), is_active=True)
        self.plan = DailyPlan.objects.create(
            date=dt.date(2026, 7, 25), work_center=self.wc, subprocess=self.sp,
            headcount=5, shift=self.shift)
        self.model = PlanningModel.objects.create(name="DAY Model")
        self.hp = HourlyPlan.objects.create(
            daily_plan=self.plan, hour=dt.time(8),
            model=self.model, planned_quantity=30)
        HourlyExecution.objects.create(hourly_plan=self.hp, actual_quantity=18)

    def test_units_convert_to_pieces(self):
        from analytics.day_service import build_day_report
        rep = build_day_report(self.plan)
        # 30 units / 3 = 10 pieces planned; 18 / 3 = 6 pieces actual.
        self.assertEqual(rep["totals"]["planned_pieces"], 10.0)
        self.assertEqual(rep["totals"]["actual_pieces"], 6.0)
        self.assertEqual(rep["totals"]["completion_pct"], 60.0)
        self.assertEqual(rep["totals"]["difference_pieces"], -4.0)

    def test_hour_classification(self):
        from analytics.day_service import classify_hour, CLASS_ACHIEVED, CLASS_NOT_MET, CLASS_EXCEEDED
        self.assertEqual(classify_hour(10, 10), CLASS_ACHIEVED)
        self.assertEqual(classify_hour(10, 8), CLASS_NOT_MET)
        self.assertEqual(classify_hour(10, 12), CLASS_EXCEEDED)

    def test_day_dashboard_renders_readonly(self):
        from django.urls import reverse
        c = Client(); c.login(username="dayuser", password="pw")
        r = c.get(reverse("analytics:day_dashboard") + f"?plan={self.plan.id}")
        self.assertEqual(r.status_code, 200)
        html = r.content.decode()
        self.assertIn("read-only", html)
        self.assertIn("Hourly Production", html)
        # No edit controls / form submission on this page.
        self.assertNotIn('type="submit"', html.split("Hourly Production")[1] if "Hourly Production" in html else "")

    def test_chart_series_in_pieces(self):
        from analytics.day_service import planned_vs_actual_series
        s = planned_vs_actual_series(self.plan)
        self.assertEqual(s["unit"], "pieces")
        self.assertEqual(s["planned"][0], 10.0)   # 30 units / 3
        self.assertEqual(s["actual"][0], 6.0)     # 18 units / 3


class DayDashboardConversionTestCase(TestCase):
    """Phase 1: DAY dashboard must apply unit→piece conversion (1/3/6) to every
    total, model summary, and chart series, and auto-classify each hour."""

    def _make_plan(self, units_per_piece):
        import random
        s = str(random.randint(10000, 99999))
        u = User.objects.create_user("dc" + s, password="pw")
        wc = WorkCenter.objects.create(name="WC-DC" + s)
        spt = SubProcessType.objects.create(
            name="T-DC" + s, applies_to="reactores", units_per_piece=units_per_piece)
        sp = SubProcess.objects.create(
            work_center=wc, name="SP-DC" + s, subprocess_type=spt)
        sh = Shift.objects.create(
            name="S-DC" + s, code="D" + s[:3],
            start_time=dt.time(6), end_time=dt.time(14), is_active=True)
        plan = DailyPlan.objects.create(
            date=dt.date(2026, 7, 21), work_center=wc, subprocess=sp,
            headcount=5, shift=sh, operator=u, created_by=u)
        return plan

    def test_conversion_factor_three(self):
        from analytics import day_service
        plan = self._make_plan(3)
        m = PlanningModel.objects.create(name="DC3 Model")
        hp = HourlyPlan.objects.create(
            daily_plan=plan, hour=dt.time(8), model=m, planned_quantity=30)
        HourlyExecution.objects.create(hourly_plan=hp, actual_quantity=18)

        rep = day_service.build_day_report(plan)
        self.assertEqual(rep["totals"]["planned_pieces"], 10)   # 30 / 3
        self.assertEqual(rep["totals"]["actual_pieces"], 6)     # 18 / 3
        self.assertEqual(rep["chart"]["planned"][0], 10)
        self.assertEqual(rep["chart"]["actual"][0], 6)
        self.assertEqual(rep["model_summary"][0]["productivity_pct"], 60.0)

    def test_hour_classification(self):
        from analytics import day_service
        self.assertEqual(day_service.classify_hour(10, 10), "Plan Achieved")
        self.assertEqual(day_service.classify_hour(10, 8),  "Plan Not Achieved")
        self.assertEqual(day_service.classify_hour(10, 12), "Production Exceeded Plan")
        self.assertEqual(day_service.classify_hour(0, 0),   "Non-Productive Hour")

    def test_twelve_hour_labels_and_completion(self):
        from analytics import day_service
        plan = self._make_plan(1)
        m = PlanningModel.objects.create(name="DC1 Model")
        hp = HourlyPlan.objects.create(
            daily_plan=plan, hour=dt.time(14), model=m, planned_quantity=10)
        HourlyExecution.objects.create(hourly_plan=hp, actual_quantity=5)
        rep = day_service.build_day_report(plan)
        self.assertEqual(rep["hourly_rows"][0]["hour_12h"], "2:00 PM")
        self.assertEqual(rep["totals"]["completion_pct"], 50.0)


class ExtendedKpiTestCase(TestCase):
    """Phase 2: productivity ranking, standard time / rate, trends, efficiency —
    all piece-aware (unit→piece conversion applied)."""

    def setUp(self):
        self.u = User.objects.create_user("kpiuser", password="pw")
        self.wc = WorkCenter.objects.create(name="WC-KPI")
        # 3 units = 1 piece.
        spt = SubProcessType.objects.create(
            name="T-KPI", applies_to="reactores", units_per_piece=3)
        self.sp = SubProcess.objects.create(
            work_center=self.wc, name="SP-KPI", subprocess_type=spt)
        self.shift = Shift.objects.create(
            name="KPI", code="KPI",
            start_time=dt.time(6), end_time=dt.time(14), is_active=True)
        self.plan = DailyPlan.objects.create(
            date=dt.date(2026, 7, 26), work_center=self.wc, subprocess=self.sp,
            headcount=5, shift=self.shift, operator=self.u, created_by=self.u)
        self.model = PlanningModel.objects.create(name="KPI Model")
        hp = HourlyPlan.objects.create(
            daily_plan=self.plan, hour=dt.time(8),
            model=self.model, planned_quantity=30)   # 10 pieces
        HourlyExecution.objects.create(hourly_plan=hp, actual_quantity=18)  # 6 pieces

    def test_production_efficiency_in_units(self):
        # Efficiency is unit-based: 18/30 = 60%.
        self.assertEqual(services.production_efficiency_pct(work_center_id=self.wc.id), 60.0)

    def test_standard_time_and_rate_piece_based(self):
        st = services.standard_time_and_rate(work_center_id=self.wc.id)
        # 6 pieces in one 60-min hour → 10 min/piece, 6 pieces/hour.
        self.assertEqual(st["standard_time_min_per_piece"], 10.0)
        self.assertEqual(st["production_rate_pieces_per_hour"], 6.0)

    def test_work_center_productivity_ranking(self):
        rows = services.top_work_centers(work_center_id=self.wc.id)
        self.assertEqual(rows[0]["name"], "WC-KPI")
        self.assertEqual(rows[0]["planned_pieces"], 10.0)
        self.assertEqual(rows[0]["actual_pieces"], 6.0)
        self.assertEqual(rows[0]["productivity_pct"], 60.0)

    def test_model_ranking_piece_based(self):
        rows = services.best_models(work_center_id=self.wc.id)
        self.assertEqual(rows[0]["name"], "KPI Model")
        self.assertEqual(rows[0]["actual_pieces"], 6.0)

    def test_productivity_trend_buckets(self):
        rows = services.productivity_trend(period="day", work_center_id=self.wc.id)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["productivity_pct"], 60.0)

    def test_extended_summary_keys(self):
        k = services.extended_kpi_summary(work_center_id=self.wc.id)
        for key in ["production_efficiency_pct", "headcount_utilization",
                    "standard_time_min_per_piece", "cycle_time_min_per_piece",
                    "production_rate_pieces_per_hour"]:
            self.assertIn(key, k)


class PeriodDashboardTestCase(TestCase):
    """Phase 3: Week/Month/Year aggregation with piece conversion, working
    hours, overtime, and comment distribution."""

    def setUp(self):
        self.u = User.objects.create_user("peruser", password="pw")
        self.wc = WorkCenter.objects.create(name="WC-PER")
        spt = SubProcessType.objects.create(
            name="T-PER", applies_to="reactores", units_per_piece=3)  # 3 units = 1 pc
        self.sp = SubProcess.objects.create(
            work_center=self.wc, name="SP-PER", subprocess_type=spt)
        self.shift = Shift.objects.create(
            name="PER", code="PER",
            start_time=dt.time(6), end_time=dt.time(14), is_active=True)
        self.model = PlanningModel.objects.create(name="PER Model")
        # Two days in the same ISO week: Mon 2026-07-20 and Tue 2026-07-21.
        for day in (dt.date(2026, 7, 20), dt.date(2026, 7, 21)):
            plan = DailyPlan.objects.create(
                date=day, work_center=self.wc, subprocess=self.sp,
                headcount=5, shift=self.shift, operator=self.u, created_by=self.u)
            hp = HourlyPlan.objects.create(
                daily_plan=plan, hour=dt.time(8), model=self.model, planned_quantity=30)
            HourlyExecution.objects.create(hourly_plan=hp, actual_quantity=18)

    def test_week_aggregates_two_days_in_pieces(self):
        from analytics import period_service as ps
        rep = ps.build_period_report("week", dt.date(2026, 7, 20), work_center_id=self.wc.id)
        # 2 days × (30 units planned, 18 actual) → 60/18? pieces: 60/3=20, 36/3=12.
        self.assertEqual(rep["totals"]["planned_pieces"], 20.0)
        self.assertEqual(rep["totals"]["actual_pieces"], 12.0)
        self.assertEqual(rep["totals"]["completion_pct"], 60.0)
        self.assertEqual(rep["working_hours"]["scheduled_hours"], 2)

    def test_month_range_and_label(self):
        from analytics import period_service as ps
        start, end, label = ps.resolve_range("month", dt.date(2026, 7, 15))
        self.assertEqual(start, dt.date(2026, 7, 1))
        self.assertEqual(end, dt.date(2026, 7, 31))
        self.assertIn("July", label)

    def test_year_aggregation(self):
        from analytics import period_service as ps
        rep = ps.build_period_report("year", dt.date(2026, 6, 1), work_center_id=self.wc.id)
        self.assertEqual(rep["totals"]["planned_pieces"], 20.0)   # same 2 rows, within year
        self.assertEqual(rep["meta"]["label"], "2026")

    def test_comment_distribution_present(self):
        from analytics import period_service as ps
        rep = ps.build_period_report("week", dt.date(2026, 7, 20), work_center_id=self.wc.id)
        types = {c["type"] for c in rep["comment_distribution"]}
        self.assertIn("Plan Not Achieved", types)   # 18 < 30 both days

    def test_week_is_iso_monday_to_sunday(self):
        from analytics import period_service as ps
        start, end, _ = ps.resolve_range("week", dt.date(2026, 7, 22))  # a Wednesday
        self.assertEqual(start.weekday(), 0)   # Monday
        self.assertEqual(end.weekday(), 6)     # Sunday


class ExportTestCase(TestCase):
    """Phase 4: Excel and PDF exports generate valid files for Day and Period."""

    def setUp(self):
        self.u = User.objects.create_user("expuser", password="pw")
        self.wc = WorkCenter.objects.create(name="WC-EXP")
        spt = SubProcessType.objects.create(
            name="T-EXP", applies_to="reactores", units_per_piece=3)
        self.sp = SubProcess.objects.create(
            work_center=self.wc, name="SP-EXP", subprocess_type=spt)
        self.shift = Shift.objects.create(
            name="EXP", code="EXP",
            start_time=dt.time(6), end_time=dt.time(14), is_active=True)
        self.plan = DailyPlan.objects.create(
            date=dt.date(2026, 7, 28), work_center=self.wc, subprocess=self.sp,
            headcount=5, shift=self.shift, operator=self.u, created_by=self.u)
        self.model = PlanningModel.objects.create(name="EXP Model")
        hp = HourlyPlan.objects.create(
            daily_plan=self.plan, hour=dt.time(8), model=self.model, planned_quantity=30)
        HourlyExecution.objects.create(hourly_plan=hp, actual_quantity=18)
        from django.test import Client
        self.client = Client()
        self.client.login(username="expuser", password="pw")

    def test_day_excel_export(self):
        from analytics import day_service, export_service
        rep = day_service.build_day_report(self.plan)
        content = export_service.day_report_to_excel(rep)
        self.assertTrue(content[:2] == b"PK")   # xlsx = zip
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content))
        self.assertEqual(wb.active.title, "Day Report")

    def test_day_pdf_export(self):
        from analytics import day_service, export_service
        rep = day_service.build_day_report(self.plan)
        content = export_service.day_report_to_pdf(rep)
        self.assertTrue(content[:5] == b"%PDF-")

    def test_period_exports(self):
        from analytics import period_service, export_service
        rep = period_service.build_period_report("month", dt.date(2026, 7, 15),
                                                 work_center_id=self.wc.id)
        self.assertTrue(export_service.period_report_to_excel(rep)[:2] == b"PK")
        self.assertTrue(export_service.period_report_to_pdf(rep)[:5] == b"%PDF-")

    def test_day_export_view_returns_file(self):
        r = self.client.get(f"/analytics/export/day/excel/?plan={self.plan.id}")
        self.assertEqual(r.status_code, 200)
        self.assertIn("attachment", r["Content-Disposition"])

    def test_period_export_view_returns_file(self):
        r = self.client.get("/analytics/export/period/pdf/?period=month&anchor=2026-07-15")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/pdf")


class PendingActualsTestCase(TestCase):
    """An hour with no execution captured must read as PENDING, not as a
    missed plan — a captured zero is a different thing entirely."""

    def setUp(self):
        import random
        s = str(random.randint(10000, 99999))
        self.wc = WorkCenter.objects.create(name="WC-PEND" + s)
        spt = SubProcessType.objects.create(
            name="T-PEND" + s, applies_to="reactores", units_per_piece=1)
        self.sp = SubProcess.objects.create(
            work_center=self.wc, name="SP-PEND" + s, subprocess_type=spt)
        self.shift = Shift.objects.create(
            name="PEND" + s, code="P" + s[:3],
            start_time=dt.time(6), end_time=dt.time(14), is_active=True)
        self.plan = DailyPlan.objects.create(
            date=dt.date(2026, 11, 5), work_center=self.wc,
            subprocess=self.sp, headcount=5, shift=self.shift)
        self.model = PlanningModel.objects.create(name="PEND Model" + s)

    def _hour(self, h, planned=10):
        return HourlyPlan.objects.create(
            daily_plan=self.plan, hour=dt.time(h),
            model=self.model, planned_quantity=planned)

    def test_uncaptured_hour_is_pending(self):
        from analytics import day_service
        self._hour(8)                      # no execution at all
        rep = day_service.build_day_report(self.plan)
        row = rep["hourly_rows"][0]
        self.assertTrue(row["pending"])
        self.assertIsNone(row["actual_pieces"])
        self.assertEqual(row["hour_status"], day_service.COMMENT_PENDING)

    def test_captured_zero_is_not_pending(self):
        from analytics import day_service
        hp = self._hour(9)
        HourlyExecution.objects.create(hourly_plan=hp, actual_quantity=0)
        rep = day_service.build_day_report(self.plan)
        row = rep["hourly_rows"][0]
        self.assertFalse(row["pending"])
        self.assertEqual(row["actual_pieces"], 0)
        self.assertEqual(row["hour_status"], day_service.COMMENT_NOT_ACHIEVED)

    def test_pending_hours_counted_in_totals(self):
        from analytics import day_service
        self._hour(8)                       # pending
        hp = self._hour(9)
        HourlyExecution.objects.create(hourly_plan=hp, actual_quantity=8)
        rep = day_service.build_day_report(self.plan)
        self.assertEqual(rep["totals"]["pending_hours"], 1)
        # Pending hours must not drag the actual total down to a false zero.
        self.assertEqual(rep["totals"]["actual_pieces"], 8)

    def test_non_productive_hour_is_not_pending(self):
        from analytics import day_service
        self._hour(8, planned=0)            # planned 0 → non-productive
        rep = day_service.build_day_report(self.plan)
        row = rep["hourly_rows"][0]
        self.assertFalse(row["pending"])
        self.assertEqual(row["hour_status"], day_service.COMMENT_NON_PRODUCTIVE)

    def test_classify_hour_pending_signature(self):
        from analytics.day_service import classify_hour, COMMENT_PENDING
        self.assertEqual(classify_hour(10, None), COMMENT_PENDING)
        self.assertEqual(classify_hour(0, None), "Non-Productive Hour")


class ScrapTotalsTestCase(TestCase):
    """Scrap must be totalled and surfaced in Analytics, Day and Period —
    and must appear in the Excel/PDF exports."""

    def setUp(self):
        import random
        s = str(random.randint(10000, 99999))
        self.wc = WorkCenter.objects.create(name="WC-SCR" + s)
        spt = SubProcessType.objects.create(
            name="T-SCR" + s, applies_to="reactores", units_per_piece=1)
        self.sp = SubProcess.objects.create(
            work_center=self.wc, name="SP-SCR" + s, subprocess_type=spt)
        self.shift = Shift.objects.create(
            name="SCR" + s, code="R" + s[:3],
            start_time=dt.time(6), end_time=dt.time(14), is_active=True)
        self.plan = DailyPlan.objects.create(
            date=dt.date(2026, 12, 12), work_center=self.wc,
            subprocess=self.sp, headcount=5, shift=self.shift)
        model = PlanningModel.objects.create(name="SCR Model" + s)
        hp1 = HourlyPlan.objects.create(daily_plan=self.plan, hour=dt.time(8),
                                        model=model, planned_quantity=10)
        hp2 = HourlyPlan.objects.create(daily_plan=self.plan, hour=dt.time(9),
                                        model=model, planned_quantity=10)
        HourlyExecution.objects.create(hourly_plan=hp1, actual_quantity=9,
                                       scrap_quantity=7)
        HourlyExecution.objects.create(hourly_plan=hp2, actual_quantity=8,
                                       scrap_quantity=5)

    def test_day_report_totals_scrap(self):
        from analytics import day_service
        rep = day_service.build_day_report(self.plan)
        self.assertEqual(rep["totals"]["scrap_units"], 12)   # 7 + 5

    def test_period_report_totals_scrap(self):
        from analytics import period_service as ps
        rep = ps.build_period_report("month", dt.date(2026, 12, 12),
                                     work_center_id=self.wc.id)
        self.assertEqual(rep["totals"]["scrap_units"], 12)

    def test_analytics_kpi_includes_scrap(self):
        k = services.kpi_summary(work_center_id=self.wc.id)
        self.assertEqual(k["scrap_quantity"], 12)

    def test_scrap_respects_subprocess_filter(self):
        """Scrap must be isolated per subprocess, not summed globally."""
        import random
        s2 = str(random.randint(10000, 99999))
        spt2 = SubProcessType.objects.create(
            name="T2-SCR" + s2, applies_to="reactores", units_per_piece=1)
        other_sp = SubProcess.objects.create(
            work_center=self.wc, name="OTHER" + s2, subprocess_type=spt2)
        other_plan = DailyPlan.objects.create(
            date=dt.date(2026, 12, 12), work_center=self.wc,
            subprocess=other_sp, headcount=5, shift=self.shift)
        m2 = PlanningModel.objects.create(name="Other Model" + s2)
        hp = HourlyPlan.objects.create(daily_plan=other_plan, hour=dt.time(8),
                                       model=m2, planned_quantity=10)
        HourlyExecution.objects.create(hourly_plan=hp, actual_quantity=5,
                                       scrap_quantity=99)
        # Filtered to the original subprocess → the 99 must NOT be included.
        k = services.kpi_summary(subprocess_id=self.sp.id)
        self.assertEqual(k["scrap_quantity"], 12)

    def test_exports_contain_scrap(self):
        import io
        from openpyxl import load_workbook
        from analytics import day_service, period_service, export_service

        day_rep = day_service.build_day_report(self.plan)
        wb = load_workbook(io.BytesIO(export_service.day_report_to_excel(day_rep)))
        labels = [str(r[0].value) for r in wb.active.iter_rows(min_col=1, max_col=1)
                  if r[0].value]
        self.assertTrue(any("Scrap" in l for l in labels))

        per_rep = period_service.build_period_report(
            "month", dt.date(2026, 12, 12), work_center_id=self.wc.id)
        wb2 = load_workbook(io.BytesIO(export_service.period_report_to_excel(per_rep)))
        labels2 = [str(r[0].value) for r in wb2.active.iter_rows(min_col=1, max_col=1)
                   if r[0].value]
        self.assertTrue(any("Scrap" in l for l in labels2))

        # Both PDFs must still build cleanly with the added column.
        self.assertTrue(export_service.day_report_to_pdf(day_rep)[:5] == b"%PDF-")
        self.assertTrue(export_service.period_report_to_pdf(per_rep)[:5] == b"%PDF-")


class DayWorkingTimeTestCase(TestCase):
    """The Day report must expose a working_time block (scheduled, real
    working, break, lunch, overtime) and the model summary must carry the
    productivity field the template actually reads."""

    def setUp(self):
        import random
        s = str(random.randint(10000, 99999))
        self.u = User.objects.create_user("wt" + s, password="pw")
        self.wc = WorkCenter.objects.create(name="WC-WT" + s)
        spt = SubProcessType.objects.create(
            name="T-WT" + s, applies_to="reactores", units_per_piece=3)
        self.sp = SubProcess.objects.create(
            work_center=self.wc, name="SP-WT" + s, subprocess_type=spt)
        self.shift = Shift.objects.create(
            name="WT" + s, code="W" + s[:3],
            start_time=dt.time(6), end_time=dt.time(14), is_active=True)
        self.plan = DailyPlan.objects.create(
            date=dt.date(2027, 1, 20), work_center=self.wc,
            subprocess=self.sp, headcount=5, shift=self.shift, created_by=self.u)
        m = PlanningModel.objects.create(name="WT Model" + s)
        h1 = HourlyPlan.objects.create(daily_plan=self.plan, hour=dt.time(8),
                                       model=m, planned_quantity=30)
        h2 = HourlyPlan.objects.create(daily_plan=self.plan, hour=dt.time(15),
                                       model=m, planned_quantity=15,
                                       is_overtime=True)
        HourlyExecution.objects.create(hourly_plan=h1, actual_quantity=18)
        HourlyExecution.objects.create(hourly_plan=h2, actual_quantity=15)
        # Time blocks auto-generate the operational events.
        HourlyPlanBlock.objects.create(daily_plan=self.plan, slot_time=dt.time(8),
                                       block_type="lunch", minutes=30,
                                       created_by=self.u)
        HourlyPlanBlock.objects.create(daily_plan=self.plan, slot_time=dt.time(15),
                                       block_type="chair", minutes=5,
                                       created_by=self.u)

    def test_working_time_block_exists(self):
        from analytics import day_service
        rep = day_service.build_day_report(self.plan)
        self.assertIn("working_time", rep)

    def test_scheduled_and_overtime_hours(self):
        from analytics import day_service
        wt = day_service.build_day_report(self.plan)["working_time"]
        self.assertEqual(wt["scheduled_hours"], 2)   # 08:00 + 15:00
        self.assertEqual(wt["overtime_hours"], 1)    # only 15:00

    def test_break_and_lunch_split_out(self):
        from analytics import day_service
        wt = day_service.build_day_report(self.plan)["working_time"]
        self.assertEqual(wt["lunch_minutes"], 30)
        self.assertEqual(wt["break_minutes"], 5)     # chair time

    def test_real_working_hours_subtracts_downtime(self):
        from analytics import day_service
        wt = day_service.build_day_report(self.plan)["working_time"]
        # 2 h scheduled = 120 min, minus 35 min downtime = 85 min = 1.4 h
        self.assertEqual(wt["downtime_minutes"], 35)
        self.assertEqual(wt["real_working_hours"], round(85 / 60, 1))

    def test_model_summary_has_productivity_pct(self):
        """The template reads m.productivity_pct — it must exist and be set."""
        from analytics import day_service
        rep = day_service.build_day_report(self.plan)
        row = rep["model_summary"][0]
        self.assertIn("productivity_pct", row)
        self.assertIsNotNone(row["productivity_pct"])

    def test_day_page_renders_working_time_values(self):
        """End-to-end: the five cards must show numbers, not blanks."""
        import re
        from django.test import Client
        c = Client(); c.force_login(self.u)
        html = c.get(f"/analytics/day/?plan={self.plan.id}").content.decode()
        for label in ["Scheduled Hrs", "Real Working Hrs", "Break",
                      "Lunch", "Overtime Hrs"]:
            i = html.find(">\n      " + label)
            self.assertNotEqual(i, -1, f"{label} card missing")
            seg = html[i:i + 800]
            m = re.search(
                r'text-2xl font-bold text-slate-800 mono">\s*(.+?)\s*</span>',
                seg, re.S)
            self.assertIsNotNone(m, f"{label} has no value node")
            self.assertNotEqual(m.group(1).strip(), "—",
                                f"{label} rendered blank")


class OvertimeBlockBeforeModelTestCase(TestCase):
    """A Time Block dropped on an overtime slot BEFORE its model is saved must
    still reach the dashboard once the hour is finally created."""

    def setUp(self):
        import random
        s = str(random.randint(10000, 99999))
        self.u = User.objects.create_user("otb" + s, password="pw")
        self.wc = WorkCenter.objects.create(name="WC-OTB" + s)
        spt = SubProcessType.objects.create(
            name="T-OTB" + s, applies_to="reactores", units_per_piece=1)
        self.sp = SubProcess.objects.create(
            work_center=self.wc, name="SP-OTB" + s, subprocess_type=spt)
        self.shift = Shift.objects.create(
            name="OTB" + s, code="O" + s[:3],
            start_time=dt.time(6), end_time=dt.time(14), is_active=True)
        self.plan = DailyPlan.objects.create(
            date=dt.date(2027, 5, 10), work_center=self.wc,
            subprocess=self.sp, headcount=5, shift=self.shift, created_by=self.u)
        self.model = PlanningModel.objects.create(name="OTB Model" + s)

    def test_block_on_hour_that_does_not_exist_yet_is_orphaned(self):
        """Before the hour exists there is nothing to attach the event to."""
        block = HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(23),
            block_type="lunch", minutes=30, created_by=self.u)
        self.assertFalse(
            ExecutionEvent.objects.filter(source_block=block).exists())

    def test_creating_the_hour_adopts_the_orphan_block(self):
        """Saving the model later must generate the missing event."""
        block = HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(23),
            block_type="lunch", minutes=30, created_by=self.u)
        HourlyPlan.objects.create(
            daily_plan=self.plan, hour=dt.time(23), model=self.model,
            planned_quantity=20, is_overtime=True)

        ev = ExecutionEvent.objects.filter(source_block=block).first()
        self.assertIsNotNone(ev, "orphan block was never adopted")
        self.assertEqual(ev.event_type.name, "Lunch")
        self.assertEqual(ev.duration_minutes, 30)

    def test_adopted_block_reaches_the_day_dashboard(self):
        from analytics import day_service
        HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(23),
            block_type="lunch", minutes=30, created_by=self.u)
        HourlyPlan.objects.create(
            daily_plan=self.plan, hour=dt.time(23), model=self.model,
            planned_quantity=20, is_overtime=True)

        wt = day_service.build_day_report(self.plan)["working_time"]
        self.assertEqual(wt["lunch_minutes"], 30)
        self.assertEqual(wt["downtime_minutes"], 30)
        self.assertEqual(wt["overtime_hours"], 1)

    def test_adopted_block_counts_as_planned_downtime(self):
        HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(23),
            block_type="lunch", minutes=30, created_by=self.u)
        HourlyPlan.objects.create(
            daily_plan=self.plan, hour=dt.time(23), model=self.model,
            planned_quantity=20, is_overtime=True)

        k = services.kpi_summary(work_center_id=self.wc.id)
        self.assertEqual(k["planned_downtime_minutes"], 30)
        self.assertEqual(k["number_of_events"], 1)

    def test_multiple_orphan_blocks_all_adopted(self):
        for btype, mins in (("lunch", 30), ("chair", 5)):
            HourlyPlanBlock.objects.create(
                daily_plan=self.plan, slot_time=dt.time(23),
                block_type=btype, minutes=mins, created_by=self.u)
        HourlyPlan.objects.create(
            daily_plan=self.plan, hour=dt.time(23), model=self.model,
            planned_quantity=20, is_overtime=True)

        events = ExecutionEvent.objects.filter(
            source_block__daily_plan=self.plan,
            source_block__slot_time=dt.time(23))
        self.assertEqual(events.count(), 2)

    def test_existing_behaviour_unchanged_when_hour_exists_first(self):
        """The original flow (model saved first) must keep working."""
        HourlyPlan.objects.create(
            daily_plan=self.plan, hour=dt.time(22), model=self.model,
            planned_quantity=15, is_overtime=True)
        block = HourlyPlanBlock.objects.create(
            daily_plan=self.plan, slot_time=dt.time(22),
            block_type="chair", minutes=5, created_by=self.u)
        self.assertTrue(
            ExecutionEvent.objects.filter(source_block=block).exists())


class RecentBoardsPanelTestCase(TestCase):
    """Recent Boards: 20 by default, +5 per 'Show more', never fewer than 20,
    and the panel can be collapsed to widen the report."""

    def setUp(self):
        import random
        s = str(random.randint(10000, 99999))
        self.u = User.objects.create_user("rb" + s, password="pw")
        wc = WorkCenter.objects.create(name="WC-RB" + s)
        spt = SubProcessType.objects.create(
            name="T-RB" + s, applies_to="reactores", units_per_piece=1)
        sp = SubProcess.objects.create(
            work_center=wc, name="SP-RB" + s, subprocess_type=spt)
        shift = Shift.objects.create(
            name="RB" + s, code="B" + s[:3],
            start_time=dt.time(6), end_time=dt.time(14), is_active=True)
        # 26 plans so pagination actually has something to page through.
        for i in range(26):
            DailyPlan.objects.create(
                date=dt.date(2027, 8, 1) + dt.timedelta(days=i),
                work_center=wc, subprocess=sp, headcount=5, shift=shift)
        from django.test import Client
        self.client = Client()
        self.client.force_login(self.u)

    def test_default_limit_is_20(self):
        r = self.client.get("/analytics/day/")
        self.assertEqual(len(r.context["plans"]), 20)
        self.assertTrue(r.context["has_more"])

    def test_show_more_adds_five(self):
        r = self.client.get("/analytics/day/?limit=25")
        self.assertEqual(len(r.context["plans"]), 25)

    def test_limit_never_drops_below_20(self):
        """?limit=5 must not shrink the list — 20 is the floor."""
        r = self.client.get("/analytics/day/?limit=5")
        self.assertEqual(len(r.context["plans"]), 20)

    def test_invalid_limit_falls_back_to_20(self):
        r = self.client.get("/analytics/day/?limit=abc")
        self.assertEqual(len(r.context["plans"]), 20)

    def test_has_more_false_when_everything_shown(self):
        r = self.client.get("/analytics/day/?limit=100")
        self.assertFalse(r.context["has_more"])

    def test_collapse_controls_are_rendered(self):
        html = self.client.get("/analytics/day/").content.decode()
        for token in ('id="boards-toggle"', 'id="boards-body"',
                      'id="day-report"', "toggleBoardsPanel"):
            self.assertIn(token, html)


class DashboardCleanupTestCase(TestCase):
    """Smart number formatting, no duplicated KPIs, and dependent filters."""

    def setUp(self):
        import random
        s = str(random.randint(10000, 99999))
        self.u = User.objects.create_user("dc" + s, password="pw")
        self.wc = WorkCenter.objects.create(name="WC-DC" + s)
        spt = SubProcessType.objects.create(
            name="T-DC" + s, applies_to="reactores", units_per_piece=3)
        self.sp = SubProcess.objects.create(
            work_center=self.wc, name="SP-DC" + s, subprocess_type=spt)
        shift = Shift.objects.create(
            name="DC" + s, code="C" + s[:3],
            start_time=dt.time(6), end_time=dt.time(14), is_active=True)
        self.plan = DailyPlan.objects.create(
            date=dt.date(2027, 9, 1), work_center=self.wc,
            subprocess=self.sp, headcount=5, shift=shift)
        m = PlanningModel.objects.create(name="DC Model" + s)
        hp = HourlyPlan.objects.create(daily_plan=self.plan, hour=dt.time(8),
                                       model=m, planned_quantity=30)
        HourlyExecution.objects.create(hourly_plan=hp, actual_quantity=18)
        from django.test import Client
        self.client = Client()
        self.client.force_login(self.u)

    # ── Number formatting ────────────────────────────────────────────────
    def test_exact_division_returns_int(self):
        from analytics.day_service import units_to_pieces
        self.assertEqual(units_to_pieces(30, 3), 10)
        self.assertIsInstance(units_to_pieces(30, 3), int)

    def test_partial_piece_keeps_decimals(self):
        """A real remainder must NOT be hidden — 10/3 is genuinely 3.33."""
        from analytics.day_service import units_to_pieces
        self.assertEqual(units_to_pieces(10, 3), 3.33)

    def test_totals_use_the_same_rule(self):
        from analytics import day_service
        rep = day_service.build_day_report(self.plan)
        self.assertEqual(rep["totals"]["planned_pieces"], 10)
        self.assertEqual(rep["totals"]["actual_pieces"], 6)

    def test_period_totals_formatted(self):
        from analytics import period_service as ps
        rep = ps.build_period_report("month", dt.date(2027, 9, 1),
                                     work_center_id=self.wc.id)
        self.assertEqual(rep["totals"]["planned_pieces"], 10)

    # ── No duplicated KPIs ───────────────────────────────────────────────
    def test_day_table_footer_no_longer_repeats_totals(self):
        html = self.client.get(
            f"/analytics/day/?plan={self.plan.id}").content.decode()
        self.assertIn("Head count for the day", html)
        self.assertNotIn(">Totals<", html)

    def test_period_banner_no_longer_repeats_kpis(self):
        html = self.client.get(
            "/analytics/period/?period=month").content.decode()
        self.assertNotIn("tracking-widest text-slate-300", html)

    def test_kpis_grouped_under_headings(self):
        html = self.client.get(
            f"/analytics/day/?plan={self.plan.id}").content.decode()
        self.assertIn("Production</h3>", html)
        self.assertIn("Working Time</h3>", html)

    # ── Tooltips survive the redesign ────────────────────────────────────
    def test_kpi_cards_still_explain_themselves(self):
        html = self.client.get(
            f"/analytics/day/?plan={self.plan.id}").content.decode()
        self.assertGreater(html.count("cursor-help"), 5)

    # ── Dependent Work Center → Subprocess ───────────────────────────────
    def test_subprocess_options_carry_their_work_center(self):
        for url in ("/analytics/", "/analytics/period/?period=month",
                    f"/analytics/day/?plan={self.plan.id}"):
            html = self.client.get(url).content.decode()
            self.assertIn("data-wc=", html, f"{url} options not tagged")
            self.assertIn("Dependent filter: Work Center", html,
                          f"{url} missing the dependent-filter script")