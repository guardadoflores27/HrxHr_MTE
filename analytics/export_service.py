# analytics/export_service.py
# ─────────────────────────────────────────────────────────────────────────────
# Excel (openpyxl) and PDF (reportlab) export of the Day and Period reports.
#
# These read the SAME report dicts produced by day_service / period_service, so
# an exported file always matches what the user sees on screen. Values are
# report snapshots (no live formulas) — a point-in-time record.
# ─────────────────────────────────────────────────────────────────────────────

import io
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer)


# ── Shared style tokens ───────────────────────────────────────────────────────

BRAND       = "1E293B"   # slate-800
BRAND_LIGHT = "334155"
ACCENT      = "2563EB"   # blue-600
GREY_HEAD   = "F1F5F9"
FONT        = "Arial"

_thin = Side(style="thin", color="E2E8F0")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def _xl_title(ws, text, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=FONT, size=15, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill = PatternFill("solid", fgColor=BRAND)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name=FONT, size=10, color="FFFFFF")
    s.fill = PatternFill("solid", fgColor=BRAND_LIGHT)
    ws.row_dimensions[2].height = 18


def _xl_section(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=11, bold=True, color=BRAND)
    return row + 1


def _xl_table(ws, start_row, headers, rows):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color=BRAND)
        c.fill = PatternFill("solid", fgColor=GREY_HEAD)
        c.alignment = Alignment(horizontal="center")
        c.border = _BORDER
    r = start_row + 1
    for row in rows:
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.font = Font(name=FONT, size=10)
            c.border = _BORDER
            c.alignment = Alignment(
                horizontal="left" if j == 1 else "right")
        r += 1
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 20 if j == 1 else 14
    return r + 1


def _kv_rows(pairs):
    return [[k, v] for k, v in pairs]


def day_report_to_excel(report):
    wb = Workbook()
    ws = wb.active
    ws.title = "Day Report"
    m, t = report["meta"], report["totals"]
    ncols = 7

    _xl_title(ws, f"Daily Production Report — {m['work_center']}",
              f"{m['date']}  ·  {m['subprocess']}  ·  Shift: {m['shift']}  ·  {m['conversion']}", ncols)

    r = 4
    r = _xl_section(ws, r, "Totals", ncols)
    r = _xl_table(ws, r, ["Metric", "Value"], _kv_rows([
        ("Planned (units)", t["planned_units"]), ("Actual (units)", t["actual_units"]),
        ("Planned (pieces)", t["planned_pieces"]), ("Actual (pieces)", t["actual_pieces"]),
        ("Scrap (units)", t["scrap_units"]),
        ("Difference (pieces)", t["difference_pieces"]),
        ("Completion %", t["completion_pct"]), ("Head Count", t["headcount"]),
    ]))

    r = _xl_section(ws, r, "Hourly Detail", ncols)
    r = _xl_table(ws, r,
        ["Hour", "Model", "Planned (pc)", "Actual (pc)",
         "HC Plan", "HC Actual", "HC Diff", "Status", "Comment"],
        [[h["hour_12h"], h["model"], h["planned_pieces"], h["actual_pieces"],
          h.get("planned_headcount"),
          h.get("actual_headcount") if h.get("actual_headcount") is not None else "—",
          h.get("headcount_diff") if h.get("headcount_diff") is not None else "—",
          h["hour_status"], h["comment"]] for h in report["hourly_rows"]])

    r = _xl_section(ws, r, "Model Summary", ncols)
    r = _xl_table(ws, r,
        ["Model", "Planned (pc)", "Actual (pc)", "Productivity %"],
        [[mm["model"], mm["planned_pieces"], mm["actual_pieces"], mm["productivity_pct"]]
         for mm in report["model_summary"]])

    return _wb_bytes(wb)


def period_report_to_excel(report):
    wb = Workbook()
    ws = wb.active
    ws.title = "Period Report"
    m, t, w = report["meta"], report["totals"], report["working_hours"]
    ncols = 5

    _xl_title(ws, f"{m['period'].title()} Production Report",
              f"{m['label']}   ·   {m['start']} → {m['end']}", ncols)

    r = 4
    r = _xl_section(ws, r, "Totals", ncols)
    r = _xl_table(ws, r, ["Metric", "Value"], _kv_rows([
        ("Planned (units)", t["planned_units"]), ("Actual (units)", t["actual_units"]),
        ("Planned (pieces)", t["planned_pieces"]), ("Actual (pieces)", t["actual_pieces"]),
        ("Scrap (units)", t["scrap_units"]),
        ("Difference (pieces)", t["difference_pieces"]),
        ("Completion %", t["completion_pct"]),
    ]))

    r = _xl_section(ws, r, "Working Hours", ncols)
    r = _xl_table(ws, r, ["Metric", "Value"], _kv_rows([
        ("Scheduled hours", w["scheduled_hours"]),
        ("Real working hours", w["real_working_hours"]),
        ("Overtime hours", w["overtime_hours"]),
        ("Planned downtime (min)", w["planned_downtime_min"]),
        ("Unplanned downtime (min)", w["unplanned_downtime_min"]),
    ]))

    r = _xl_section(ws, r, "Comment Distribution", ncols)
    r = _xl_table(ws, r, ["Status", "Count"],
        [[c["type"], c["count"]] for c in report["comment_distribution"]])

    r = _xl_section(ws, r, "Model Summary", ncols)
    r = _xl_table(ws, r,
        ["Model", "Planned (pc)", "Actual (pc)", "Productivity %"],
        [[mm["model"], mm["planned_pieces"], mm["actual_pieces"], mm["productivity_pct"]]
         for mm in report["model_summary"]])

    return _wb_bytes(wb)


def _wb_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# PDF
# ═══════════════════════════════════════════════════════════════════════════

def _pdf_styles():
    ss = getSampleStyleSheet()
    ss.add(ParagraphStyle("BrandTitle", parent=ss["Title"], fontName="Helvetica-Bold",
                          fontSize=16, textColor=colors.white, alignment=0, leading=20))
    ss.add(ParagraphStyle("BrandSub", parent=ss["Normal"], fontName="Helvetica",
                          fontSize=9, textColor=colors.white, alignment=0))
    ss.add(ParagraphStyle("Section", parent=ss["Heading2"], fontName="Helvetica-Bold",
                          fontSize=11, textColor=colors.HexColor("#" + BRAND), spaceBefore=10, spaceAfter=4))
    return ss


def _header_band(title, subtitle, styles, width):
    band = Table([[Paragraph(title, styles["BrandTitle"])],
                  [Paragraph(subtitle, styles["BrandSub"])]], colWidths=[width])
    band.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + BRAND)),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#" + BRAND_LIGHT)),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    return band


def _data_table(headers, rows, col_widths):
    data = [headers] + rows
    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + GREY_HEAD)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#" + BRAND)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    t.setStyle(TableStyle(style))
    return t


def day_report_to_pdf(report):
    m, t = report["meta"], report["totals"]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    styles = _pdf_styles()
    width = doc.width
    story = []

    story.append(_header_band(
        f"Daily Production Report — {m['work_center']}",
        f"{m['date']}  ·  {m['subprocess']}  ·  Shift: {m['shift']}  ·  {m['conversion']}",
        styles, width))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Totals", styles["Section"]))
    story.append(_data_table(
        ["Planned (pc)", "Actual (pc)", "Scrap (u)", "Difference (pc)", "Completion %", "Head Count"],
        [[t["planned_pieces"], t["actual_pieces"], t["scrap_units"],
          t["difference_pieces"], _pct(t["completion_pct"]), t["headcount"]]],
        [width/6.0]*6))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Hourly Detail", styles["Section"]))
    hrows = [[h["hour_12h"], h["model"], h["planned_pieces"], h["actual_pieces"],
              h.get("planned_headcount"),
              h.get("actual_headcount") if h.get("actual_headcount") is not None else "—",
              h.get("headcount_diff") if h.get("headcount_diff") is not None else "—",
              h["hour_status"], (h["comment"] or "")[:32]]
             for h in report["hourly_rows"]]
    story.append(_data_table(
        ["Hour", "Model", "Planned", "Actual", "HC Plan", "HC Real", "HC Diff",
         "Status", "Comment"],
        hrows, [w*width for w in (0.10, 0.14, 0.08, 0.08, 0.07, 0.07, 0.07, 0.16, 0.23)]))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Model Summary", styles["Section"]))
    mrows = [[mm["model"], mm["planned_pieces"], mm["actual_pieces"], _pct(mm["productivity_pct"])]
             for mm in report["model_summary"]]
    story.append(_data_table(
        ["Model", "Planned (pc)", "Actual (pc)", "Productivity %"],
        mrows, [width*0.4, width*0.2, width*0.2, width*0.2]))

    story.append(Spacer(1, 12))
    story.append(Paragraph(
        f"Generated {dt.datetime.now():%Y-%m-%d %H:%M} · MTE Hr×Hr",
        ParagraphStyle("foot", fontName="Helvetica-Oblique", fontSize=7,
                       textColor=colors.HexColor("#94A3B8"))))
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def period_report_to_pdf(report):
    m, t, w = report["meta"], report["totals"], report["working_hours"]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    styles = _pdf_styles()
    width = doc.width
    story = []

    story.append(_header_band(
        f"{m['period'].title()} Production Report",
        f"{m['label']}   ·   {m['start']} → {m['end']}", styles, width))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Totals", styles["Section"]))
    story.append(_data_table(
        ["Planned (pc)", "Actual (pc)", "Scrap (u)", "Difference (pc)", "Completion %"],
        [[t["planned_pieces"], t["actual_pieces"], t["scrap_units"],
          t["difference_pieces"], _pct(t["completion_pct"])]],
        [width/5.0]*5))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Working Hours", styles["Section"]))
    story.append(_data_table(
        ["Scheduled", "Real Working", "Overtime", "Planned DT (min)", "Unplanned DT (min)"],
        [[w["scheduled_hours"], w["real_working_hours"], w["overtime_hours"],
          w["planned_downtime_min"], w["unplanned_downtime_min"]]],
        [width/5.0]*5))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Comment Distribution", styles["Section"]))
    story.append(_data_table(
        ["Status", "Count"],
        [[c["type"], c["count"]] for c in report["comment_distribution"]] or [["—", 0]],
        [width*0.6, width*0.4]))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Model Summary", styles["Section"]))
    mrows = [[mm["model"], mm["planned_pieces"], mm["actual_pieces"], _pct(mm["productivity_pct"])]
             for mm in report["model_summary"]]
    story.append(_data_table(
        ["Model", "Planned (pc)", "Actual (pc)", "Productivity %"],
        mrows or [["—", 0, 0, "—"]], [width*0.4, width*0.2, width*0.2, width*0.2]))

    story.append(Spacer(1, 12))
    story.append(Paragraph(
        f"Generated {dt.datetime.now():%Y-%m-%d %H:%M} · MTE Hr×Hr",
        ParagraphStyle("foot", fontName="Helvetica-Oblique", fontSize=7,
                       textColor=colors.HexColor("#94A3B8"))))
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def _pct(v):
    return "—" if v is None else f"{v}%"